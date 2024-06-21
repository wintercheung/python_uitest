#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author  : winter cheung
# @Time    : 2024/6/21 下午1:11

from openpyxl.styles import PatternFill, Font


def write_pass(cell, row, column):
    """
    写入 pass 的配置
    :param cell:
    :param row:
    :param column:
    :return:
    """
    cell(row=row, column=column).value = "PASS"
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')  # 单元格显示：绿色加粗
    cell(row=row, column=column).font = Font(bold=True)


def write_fail(cell, row, column):
    """
    写入 failed 的配置
    :param cell:
    :param row:
    :param column:
    :return:
    """
    cell(row=row, column=column).value = 'FAILED'
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')  # 单元格显示：红色加粗
    cell(row=row, column=column).font = Font(bold=True)